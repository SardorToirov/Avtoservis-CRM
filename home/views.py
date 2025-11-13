from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from .models import Mijoz, Avto, Tashrif
from .forms import MijozForms, AvtoForms, TashrifForms
from django.db.models import Q
from openpyxl import Workbook


# Dashboard
def home(request):
    context = {
        'mijoz_soni': Mijoz.objects.count(),
        'avto_soni': Avto.objects.count(),
        'tashrif_soni': Tashrif.objects.count(),
    }
    return render(request, 'home.html', context)

# ===== Mijozlar =====
def mijoz_list(request):
    search_query = request.GET.get('q', '')
    if search_query:
        mijozlar = Mijoz.objects.filter(
            Q(ism__icontains=search_query) |
            Q(familiya__icontains=search_query) |
            Q(telefon_raqami__icontains=search_query)
        )
    else:
        mijozlar = Mijoz.objects.all()
    context = {'mijozlar': mijozlar, 'search_query': search_query}
    return render(request, 'mijoz_list.html', context)

def mijoz_create(request):
    form = MijozForms(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('mijoz_list')
    return render(request, 'mijoz_form.html', {'form': form})

def mijoz_edit(request, pk):
    mijoz = get_object_or_404(Mijoz, pk=pk)
    form = MijozForms(request.POST or None, instance=mijoz)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('mijoz_list')
    return render(request, 'mijoz_form.html', {'form': form})

def mijoz_delete(request, pk):
    mijoz = get_object_or_404(Mijoz, pk=pk)
    mijoz.delete()
    return redirect('mijoz_list')


# ===== Avtolar =====
def avto_list(request):
    search_query = request.GET.get('q', '')
    if search_query:
        avtolar = Avto.objects.filter(
            Q(model__icontains=search_query) |
            Q(raqam__icontains=search_query) |
            Q(mijoz__ism__icontains=search_query)
        )
    else:
        avtolar = Avto.objects.all()
    context = {'avtolar': avtolar, 'search_query': search_query}
    return render(request, 'avto_list.html', context)

def avto_create(request):
    form = AvtoForms(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('avto_list')
    return render(request, 'avto_form.html', {'form': form})

def avto_edit(request, pk):
    avto = get_object_or_404(Avto, pk=pk)
    form = AvtoForms(request.POST or None, instance=avto)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('avto_list')
    return render(request, 'avto_create.html', {'form': form})

def avto_delete(request, pk):
    avto = get_object_or_404(Avto, pk=pk)
    avto.delete()
    return redirect('avto_list')


# ===== Tashriflar =====
def tashrif_list(request):
    search_query = request.GET.get('q', '')
    if search_query:
        tashriflar = Tashrif.objects.filter(
            Q(avto__raqam__icontains=search_query) |
            Q(avto__model__icontains=search_query) |
            Q(avto__mijoz__ism__icontains=search_query) |
            Q(sana__icontains=search_query)
        )
    else:
        tashriflar = Tashrif.objects.all()
    context = {'tashriflar': tashriflar, 'search_query': search_query}
    return render(request, 'tashrif_list.html', context)

def tashrif_create(request):
    form = TashrifForms(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('tashrif_list')
    return render(request, 'tashrif_form.html', {'form': form})

def tashrif_edit(request, pk):
    tashrif = get_object_or_404(Tashrif, pk=pk)
    form = TashrifForms(request.POST or None, instance=tashrif)
    if request.method == 'POST' and form.is_valid():
        form.save()
        return redirect('tashrif_list')
    return render(request, 'tashrif_form.html', {'form': form})

def tashrif_delete(request, pk):
    tashrif = get_object_or_404(Tashrif, pk=pk)
    tashrif.delete()
    return redirect('tashrif_list')


from django.http import HttpResponse
from openpyxl import Workbook
from .models import Mijoz

def export_mijoz_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mijozlar"
    ws.append(["Ismi", "Familiya", "Telefon"])
    mijozlar = Mijoz.objects.all()
    for m in mijozlar:
        ws.append([m.ism, m.familiya, m.telefon_raqami])
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="mijozlar.xlsx"'
    wb.save(response)
    return response


# 🔹 Avtomobillarni Excelga eksport qilish
def export_avto_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Avtomobillar"
    ws.append(["Model", "Davlat raqami", "Mijoz"])

    avto_list = Avto.objects.all()
    for a in avto_list:
        ws.append([
            a.model,
            a.raqam,
            str(a.mijoz)

        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="avtomobillar.xlsx"'
    wb.save(response)
    return response


# 🔹 Tashriflarni Excelga eksport qilish
def export_tashrif_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tashriflar"
    ws.append(["Avto", "Sana", "Joriy km", "keyngi sana","Keyingi km", "Xizmatlar"])

    tashriflar = Tashrif.objects.all()
    for t in tashriflar:
        ws.append([
            str(t.avto),
            t.sana.strftime("%Y-%m-%d") if t.sana else "",
            t.joriy_km,
            t.keyingi_sana,
            t.keyingi_km,
            t.xizmatlar
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="tashriflar.xlsx"'
    wb.save(response)
    return response


from openpyxl import Workbook
from django.http import HttpResponse
from .models import Mijoz, Avto, Tashrif
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from .models import Mijoz, Avto, Tashrif


def export_all_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Barcha ma'lumotlar"

    # --- Sarlavhalar ---
    headers = [
        "Turi", "Ismi", "Familiyasi", "Telefon",
        "Avto modeli", "Davlat raqami",
        "Sana", "Joriy km", "Keyingi sana", "Keyingi km", "Xizmatlar"
    ]
    ws.append(headers)

    # --- Sarlavha dizayni ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="406ABF", end_color="406ABF", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 18

    # --- Jadvalga ketma-ket ma'lumot qo'shish ---
    # Mijoz -> Avto -> Tashrif ketma-ket chiqadi

    # Mijozlar
    for m in Mijoz.objects.all():
        ws.append([
            "👤 Mijoz",
            m.ism,
            m.familiya,
            m.telefon_raqami,
            "", "", "", "", "", "", ""
        ])

        # Shu mijozga tegishli avtolarni chiqarish
        avtolar = Avto.objects.filter(mijoz=m)
        for a in avtolar:
            ws.append([
                "   🚘 Avto",
                m.ism,
                m.familiya,
                m.telefon_raqami,
                a.model,
                a.raqam,
                "", "", "", "", ""
            ])

            # Shu avto uchun tashriflarni chiqarish
            tashriflar = Tashrif.objects.filter(avto=a)
            for t in tashriflar:
                ws.append([
                    "      🧾 Tashrif",
                    m.ism,
                    m.familiya,
                    m.telefon_raqami,
                    a.model,
                    a.raqam,
                    t.sana.strftime("%Y-%m-%d") if t.sana else "",
                    t.joriy_km,
                    t.keyingi_sana.strftime("%Y-%m-%d") if t.keyingi_sana else "",
                    t.keyingi_km,
                    t.xizmatlar
                ])

    # --- Hamma kataklarga chegara berish ---
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # --- Excel faylni yuborish ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="barcha_malumotlar_tartibli.xlsx"'
    wb.save(response)
    return response
